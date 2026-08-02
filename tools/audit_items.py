#!/usr/bin/env python3
"""道具引用全景审计 → 生成 Excel。

跑法: ./agent_server/.venv/bin/python tools/audit_items.py

把散落在 6 处数据源里的道具引用汇总到一张表，用来人工核查：
  items.py          道具定义（id/名字/base_value）
  data/animals/*    喜好(loves/likes)、厌恶(dislikes/hates)、
                    三档回礼池、signature_gift、forage 野货池
  world/quests.json 任务需求物 / 任务奖励物
  world/crises.json 危机选项消耗物
  world/spawners    地图刷新点（=玩家可自行捡到）
  world/milestones  里程碑奖励物
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "agent_server"))

import items as items_mod  # noqa: E402

ITEMS = items_mod._ITEMS
DATA = ROOT / "data"


def _load(p: Path):
    with p.open(encoding="utf-8") as f:
        return json.load(f)


def _strip(d: dict) -> dict:
    return {k: v for k, v in d.items() if not k.startswith("_")}


# ---------- 收集 ----------
personas = {}
for f in sorted((DATA / "animals").glob("*.json")):
    p = _load(f)
    personas[p["id"]] = p

quests = _strip(_load(DATA / "world" / "quests.json"))
crises = _strip(_load(DATA / "world" / "crises.json"))
spawn_raw = _load(DATA / "world" / "spawners.json")
milestones = _strip(_load(DATA / "world" / "milestones.json"))

# 在场名单（备选池的人也要审，但要标出来）
try:
    import roster
    PRESENT = set(roster.DEFAULT_PRESENT)
except Exception:
    PRESENT = set(personas)

NAME = {aid: p.get("name", aid) for aid, p in personas.items()}

# item_id -> 各类引用
loves = defaultdict(list)
likes = defaultdict(list)
dislikes = defaultdict(list)
hates = defaultdict(list)
ret_common = defaultdict(list)
ret_mid = defaultdict(list)
ret_rare = defaultdict(list)
signature = defaultdict(list)
forage = defaultdict(list)

for aid, p in personas.items():
    n = NAME[aid]
    gp = p.get("gift_prefs", {}) or {}
    for i in gp.get("loves", []) or []:
        loves[i].append(n)
    for i in gp.get("likes", []) or []:
        likes[i].append(n)
    for i in gp.get("dislikes", []) or []:
        dislikes[i].append(n)
    for i in gp.get("hates", []) or []:
        hates[i].append(n)
    for i in p.get("return_gifts", []) or []:
        ret_common[i].append(n)
    for i in p.get("mid_return_gifts", []) or []:
        ret_mid[i].append(n)
    for i in p.get("rare_return_gifts", []) or []:
        ret_rare[i].append(n)
    sig = p.get("signature_gift")
    if sig:
        signature[sig].append(n)
    fg = (p.get("movement", {}) or {}).get("forage", {}) or {}
    for i in fg.get("items", []) or []:
        forage[i].append(n)

quest_need = defaultdict(list)
quest_reward = defaultdict(list)
for qid, q in quests.items():
    npc = NAME.get(q.get("npc_id", ""), q.get("npc_id", ""))
    req = q.get("requires", {}) or {}
    it = req.get("item_id")
    if it:
        quest_need[it].append(f"{npc}·{q.get('title', qid)}×{req.get('count', 1)}")
    rw = q.get("reward", {}) or {}
    it2 = rw.get("item_id")
    if it2:
        quest_reward[it2].append(f"{npc}·{q.get('title', qid)}×{rw.get('count', 1)}")

crisis_cost = defaultdict(list)
for cid, c in crises.items():
    for o in c.get("options", []) or []:
        r = o.get("requires", {}) or {}
        it = r.get("item")
        if it:
            crisis_cost[it].append(f"{c.get('title', cid)}×{r.get('count', 1)}")

ground = defaultdict(int)
for s in spawn_raw.get("spawners", []) or []:
    it = s.get("item_id")
    if it:
        ground[it] += len(s.get("anchors", []) or [])

ms_reward = defaultdict(list)
for aid, entries in milestones.items():
    if not isinstance(entries, list):
        continue
    for m in entries:
        it = (m.get("reward", {}) or {}).get("item_id")
        if it:
            ms_reward[it].append(f"{NAME.get(aid, aid)}·{m.get('title', '')}")

ALL_REF = [loves, likes, dislikes, hates, ret_common, ret_mid, ret_rare,
           signature, forage, quest_need, quest_reward, crisis_cost, ms_reward]

# 被引用但没有定义的 id（配置写错时会出现）
referenced = set()
for d in ALL_REF:
    referenced |= set(d)
referenced |= set(ground)
UNDEFINED = sorted(referenced - set(ITEMS))

print(f"道具定义 {len(ITEMS)} 件 / 被引用 {len(referenced)} 个 id")
if UNDEFINED:
    print(f"⚠ 引用了未定义的 id: {UNDEFINED}")


# ---------- 逐件道具分析 ----------
def j(lst):
    return "、".join(lst) if lst else ""


def holders_of(i: str):
    """真正「手上有」这件东西的 NPC（可被索要/回赠的来源）。"""
    return sorted(set(ret_common[i]) | set(ret_mid[i]) | set(ret_rare[i])
                  | set(signature[i]) | set(forage[i]))


rows = []
issues = []

for iid, it in sorted(ITEMS.items(), key=lambda kv: (-kv[1].base_value, kv[0])):
    src = holders_of(iid)
    g = ground.get(iid, 0)
    need = quest_need[iid]
    is_ground = g > 0

    # 获取途径
    ways = []
    if is_ground:
        ways.append(f"地图可捡({g}点)")
    if src:
        ways.append(f"NPC:{j(src)}")
    if quest_reward[iid]:
        ways.append(f"任务奖励×{len(quest_reward[iid])}")
    if ms_reward[iid]:
        ways.append(f"里程碑×{len(ms_reward[iid])}")

    # 用途
    uses = []
    if loves[iid] or likes[iid]:
        uses.append(f"送礼({len(loves[iid])}爱/{len(likes[iid])}喜)")
    if need:
        uses.append(f"任务需求×{len(need)}")
    if crisis_cost[iid]:
        uses.append(f"危机消耗×{len(crisis_cost[iid])}")

    # ---- 异常判定 ----
    flags = []
    if not ways:
        flags.append("❌无获取途径")
    if not uses:
        flags.append("⚠无任何用途")
    if need and not ways:
        flags.append("❌任务要它但拿不到")
    if is_ground and it.base_value > 3:
        flags.append(f"❌地图可捡却 base_value={it.base_value}(铁律≤3)")
    # 刷好感闭环：某 NPC 既送出它、又喜欢收到它
    loop = sorted((set(ret_common[iid]) | set(ret_mid[iid]) | set(ret_rare[iid]))
                  & (set(loves[iid]) | set(likes[iid])))
    if loop:
        flags.append(f"❌刷好感闭环:{j(loop)}")
    if not loves[iid] and not likes[iid] and not dislikes[iid] and not hates[iid]:
        flags.append("⚠无人对它有偏好")

    rows.append({
        "id": iid,
        "名称": it.name,
        "价值": it.base_value,
        "获取途径": j(ways) or "（无）",
        "用途": j(uses) or "（无）",
        "爱": j(loves[iid]),
        "喜欢": j(likes[iid]),
        "不喜欢": j(dislikes[iid]),
        "讨厌": j(hates[iid]),
        "回礼-普通": j(ret_common[iid]),
        "回礼-中档": j(ret_mid[iid]),
        "回礼-稀有": j(ret_rare[iid]),
        "招牌礼": j(signature[iid]),
        "NPC野货池": j(forage[iid]),
        "地图刷新点": g or "",
        "任务需求": j(need),
        "任务奖励": j(quest_reward[iid]),
        "危机消耗": j(crisis_cost[iid]),
        "里程碑奖励": j(ms_reward[iid]),
        "问题": j(flags),
        "描述": it.desc,
    })
    if flags:
        issues.append((iid, it.name, it.base_value, j(flags)))

print(f"\n发现 {len(issues)} 件道具存在问题：")
for iid, nm, bv, f in issues:
    print(f"  {nm}({iid}) v={bv}  {f}")

# ---------- 写 Excel ----------
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

HDR_FILL = PatternFill("solid", fgColor="4472C4")
HDR_FONT = Font(color="FFFFFF", bold=True)
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")

wb = Workbook()


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def autosize(ws, cols):
    for idx, w in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# --- Sheet1 道具全景 ---
ws = wb.active
ws.title = "道具全景"
COLS = list(rows[0].keys())
ws.append(COLS)
for r in rows:
    ws.append([r[c] for c in COLS])
style_header(ws, len(COLS))
autosize(ws, [12, 10, 6, 26, 22, 16, 20, 12, 12,
              14, 14, 14, 10, 16, 10, 30, 26, 18, 18, 34, 30])
pi = COLS.index("问题") + 1
for ri in range(2, len(rows) + 2):
    v = str(ws.cell(row=ri, column=pi).value or "")
    if "❌" in v:
        for c in range(1, len(COLS) + 1):
            ws.cell(row=ri, column=c).fill = BAD_FILL
    elif "⚠" in v:
        for c in range(1, len(COLS) + 1):
            ws.cell(row=ri, column=c).fill = WARN_FILL

# --- Sheet2 异常清单 ---
ws2 = wb.create_sheet("异常清单")
ws2.append(["id", "名称", "价值", "问题", "说明"])
EXPLAIN = {
    "❌无获取途径": "玩家在游戏里根本拿不到这件东西：地图不刷、没有 NPC 会给、任务也不奖励。",
    "⚠无任何用途": "拿到了也没用：没人喜欢收、没有任务要、危机也用不上。",
    "⚠无人对它有偏好": "送给任何 NPC 都只有默认反应，没有惊喜也没有厌恶。",
    "❌任务要它但拿不到": "任务需求物没有任何来源，该任务无法完成（死锁）。",
    "❌刷好感闭环": "同一个 NPC 既会把它送出来、又喜欢收到它 → 要走再送回可无限刷好感。",
}
for iid, nm, bv, f in issues:
    exp = "；".join(v for k, v in EXPLAIN.items() if k.split(":")[0] in f)
    ws2.append([iid, nm, bv, f, exp])
style_header(ws2, 5)
autosize(ws2, [14, 12, 6, 40, 60])
for ri in range(2, len(issues) + 2):
    fill = BAD_FILL if "❌" in str(ws2.cell(row=ri, column=4).value) else WARN_FILL
    for c in range(1, 6):
        ws2.cell(row=ri, column=c).fill = fill

# --- Sheet3 NPC×道具偏好矩阵 ---
ws3 = wb.create_sheet("NPC偏好矩阵")
npc_ids = sorted(personas, key=lambda a: (a not in PRESENT, a))
hdr = ["id", "名称", "价值"] + [
    NAME[a] + ("" if a in PRESENT else "(备选)") for a in npc_ids]
ws3.append(hdr)
MARK = {"love": "爱", "like": "喜", "dislike": "不喜", "hate": "厌",
        "common": "回礼", "mid": "中礼", "rare": "稀礼",
        "sig": "招牌", "forage": "野货"}
for iid, it in sorted(ITEMS.items(), key=lambda kv: (-kv[1].base_value, kv[0])):
    line = [iid, it.name, it.base_value]
    for a in npc_ids:
        n = NAME[a]
        tags = []
        if n in loves[iid]:
            tags.append(MARK["love"])
        if n in likes[iid]:
            tags.append(MARK["like"])
        if n in dislikes[iid]:
            tags.append(MARK["dislike"])
        if n in hates[iid]:
            tags.append(MARK["hate"])
        if n in ret_common[iid]:
            tags.append(MARK["common"])
        if n in ret_mid[iid]:
            tags.append(MARK["mid"])
        if n in ret_rare[iid]:
            tags.append(MARK["rare"])
        if n in signature[iid]:
            tags.append(MARK["sig"])
        if n in forage[iid]:
            tags.append(MARK["forage"])
        line.append("/".join(tags))
    ws3.append(line)
style_header(ws3, len(hdr))
autosize(ws3, [14, 12, 6] + [14] * len(npc_ids))

# --- Sheet4 NPC 视角汇总 ---
ws4 = wb.create_sheet("NPC汇总")
ws4.append(["NPC", "状态", "爱(数)", "喜欢(数)", "不喜欢(数)", "讨厌(数)",
            "普通回礼", "中档回礼", "稀有回礼", "招牌礼", "野货池",
            "发布任务数", "任务需求物", "任务奖励物"])
for a in npc_ids:
    p = personas[a]
    gp = p.get("gift_prefs", {}) or {}
    qs = [q for q in quests.values() if q.get("npc_id") == a]
    ws4.append([
        NAME[a], "在场" if a in PRESENT else "备选池",
        len(gp.get("loves", []) or []), len(gp.get("likes", []) or []),
        len(gp.get("dislikes", []) or []), len(gp.get("hates", []) or []),
        j([ITEMS[i].name for i in (p.get("return_gifts") or []) if i in ITEMS]),
        j([ITEMS[i].name for i in (p.get("mid_return_gifts") or []) if i in ITEMS]),
        j([ITEMS[i].name for i in (p.get("rare_return_gifts") or []) if i in ITEMS]),
        ITEMS[p["signature_gift"]].name if p.get("signature_gift") in ITEMS else "",
        j([ITEMS[i].name for i in
           ((p.get("movement", {}) or {}).get("forage", {}) or {}).get("items", [])
           if i in ITEMS]),
        len(qs),
        j(sorted({ITEMS[q["requires"]["item_id"]].name for q in qs
                  if (q.get("requires") or {}).get("item_id") in ITEMS})),
        j(sorted({ITEMS[q["reward"]["item_id"]].name for q in qs
                  if (q.get("reward") or {}).get("item_id") in ITEMS})),
    ])
style_header(ws4, 14)
autosize(ws4, [12, 10, 8, 10, 11, 10, 26, 22, 22, 12, 30, 11, 34, 30])

# --- Sheet5 总览 ---
ws5 = wb.create_sheet("总览", 0)
ws5.append(["项目", "数值", "备注"])
ground_items = [i for i in ITEMS if ground.get(i)]
no_way = [i for i in ITEMS if not holders_of(i) and not ground.get(i)
          and not quest_reward[i] and not ms_reward[i]]
no_use = [i for i in ITEMS if not (loves[i] or likes[i] or quest_need[i]
                                   or crisis_cost[i])]
stats = [
    ("道具总数", len(ITEMS), "items.py 中定义"),
    ("在场 NPC", len(PRESENT), "备选池另有 %d 人" % (len(personas) - len(PRESENT))),
    ("任务总数", len(quests), "quests.json"),
    ("危机总数", len(crises), "crises.json"),
    ("地图可捡道具", len(ground_items), "有 spawner 刷新点"),
    ("无获取途径", len(no_way), "❌ 玩家拿不到：" + j([ITEMS[i].name for i in no_way])),
    ("无任何用途", len(no_use), "⚠ 拿到没用：" + j([ITEMS[i].name for i in no_use])),
    ("存在问题的道具", len(issues), "见「异常清单」页"),
]
for s in stats:
    ws5.append(list(s))
style_header(ws5, 3)
autosize(ws5, [18, 10, 80])

OUT = ROOT / "docs" / "道具审计表.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"\n已生成: {OUT}")
print(f"  总览 / 道具全景({len(rows)}行) / 异常清单({len(issues)}行) / "
      f"NPC偏好矩阵 / NPC汇总({len(npc_ids)}人)")




